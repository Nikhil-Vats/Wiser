from __future__ import unicode_literals
from django.shortcuts import render
from django.views.generic import View,ListView
from django.http import JsonResponse,HttpResponseRedirect,HttpResponse
from main.models import Userdata,Question,Answer
import json
from django.contrib.auth.models import User
import datetime as dt
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
import openpyxl
from openpyxl.utils import get_column_letter

def main(request):
    return render(request, 'index.html',)

@login_required(login_url='/')
def proceed(request):
    
    return render(request,'loggedIn.html',)

@login_required(login_url='/')
def formdata(request):
    
    data_get = json.loads(request.body.decode('utf-8'))
    Userdata.objects.create(user_id = request.user,
                            name=data_get['name'],
                            age=data_get['age'],
                            gender=data_get['gender'],
                            contactno=data_get['phone'],
                            education_level=data_get['year'],
                            state=data_get['state'],
                            city=data_get['city'])
    return JsonResponse({"success":1})


@login_required(login_url='/')
def countries_list(request):
    data = [];
    for i in Country.objects.all();
        data.append({"pk":i.pk,"name":i.name})
        
    return JsonResponse({"data":data})

@login_required(login_url='/')
def state_list(request):
    data_get = json.loads(request.body.decode('utf-8'))
    data = []
    for i in Country.objects.get(pk=data_get['pk'].state_set.all():
         data.append({"pk":i.pk,"name":i.name})
                                 
    return JsonResponse({"data":data})


@login_required(login_url='/')
def pre_cat(request):
    
    data_get = json.loads(request.body.decode('utf-8'))
    ud = Userdata.objects.get(user_id = request.user)
    ud.category = data_get['category']
    ud.save()
    return JsonResponse({"success":1})    

@login_required(login_url='/')
def status(request):
    
    try:
        ud = Userdata.objects.get(user_id = request.user)
        st = ud.status
        cat = ud.category
    except(Userdata.DoesNotExist,TypeError,ValueError,OverflowError):
        st = 0
        cat = -1

    data = {"stage":st,"category":cat}
    return JsonResponse(data)
    
@login_required(login_url='/')
def prepos_details(request):
    
    if request.method=='POST':
        ud = Userdata.objects.get(user_id = request.user)
        dat = []
        qlist = Question.objects.none()
        r = 0
        eqno = -1

        if ud.status == 1: 
            qlist = Question.objects.filter(q_category=ud.category,q_type='prepos')

            for q in qlist:
                if q.q_format=="mcq":
                    dat.append({"format":"mcq","pk":q.pk,"text":q.text,"choice1":q.choice1,"choice2":q.choice2,"choice3":q.choice3,"choice4":q.choice4,"choice5":q.choice5,"choice6":q.choice6,"choice7":q.choice7})
                elif q.q_format=="openended":
                    dat.append({"format":"openended","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2})
                else:
                    dat.append({"format":"radio","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2})


        elif ud.status == 2:
            if ud.email_date <= dt.date.today():
                if ud.category != 1:
                    qlist = Question.objects.filter(q_category=ud.category-1,q_type='experiment',q_format="mcq")
                    r = qlist.count()

                    for q in qlist:
                        dat.append({"format":"mcq","pk":q.pk,"text":q.text,"choice1":q.choice1,"choice2":q.choice2,"choice3":q.choice3,"choice4":q.choice4,"choice5":q.choice5,"choice6":q.choice6,"choice7":q.choice7})                   


                qlist = Question.objects.filter(q_category=ud.category,q_type='experiment',q_format="mcq")

                for q in qlist:
                    dat.append({"format":"mcq","pk":q.pk,"text":q.text,"choice1":q.choice1,"choice2":q.choice2,"choice3":q.choice3,"choice4":q.choice4,"choice5":q.choice5,"choice6":q.choice6,"choice7":q.choice7})                   


                qlist = Question.objects.filter(q_category=ud.category,q_type='experiment',q_format="radio")

                for q in qlist:
                    dat.append({"format":"radio","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2})

                eqno = len(dat)

                qlist = Question.objects.filter(q_category=ud.category,q_type='experiment',q_format="openended")

                for q in qlist:
                    dat.append({"format":"openended","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2,"download":q.choice3})
                    
            else:
                return JsonResponse({"data":-1,"qno":-1,"totq":-1,"rqno":-1})

        elif ud.status == 3:
            if ud.email_date<=dt.date.today():
                qlist = Question.objects.filter(q_category=ud.category,q_type='placebo')
                eqno = qlist.count()

                for q in qlist:
                    if q.q_format=="mcq":
                        dat.append({"format":"mcq","pk":q.pk,"text":q.text,"choice1":q.choice1,"choice2":q.choice2,"choice3":q.choice3,"choice4":q.choice4,"choice5":q.choice5,"choice6":q.choice6,"choice7":q.choice7})
                    elif q.q_format=="openended":
                        dat.append({"format":"openended","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2})
                    else:
                        dat.append({"format":"radio","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2})

            else:
                return JsonResponse({"data":-1,"qno":-1,"totq":-1,"rqno":-1})

        elif ud.status == 5:
            if ud.email_date<=dt.date.today():
                if ud.category == 1:
                    qlist = Question.objects.filter(q_category=4,q_type='experiment',q_format="mcq")
                    r = qlist.count()

                    for q in qlist:
                        dat.append({"format":"mcq","pk":q.pk,"text":q.text,"choice1":q.choice1,"choice2":q.choice2,"choice3":q.choice3,"choice4":q.choice4,"choice5":q.choice5,"choice6":q.choice6,"choice7":q.choice7})
                        


                qlist = Question.objects.filter(q_category=ud.category,q_type='prepos')

                for q in qlist:
                    if q.q_format=="mcq":
                        dat.append({"format":"mcq","pk":q.pk,"text":q.text,"choice1":q.choice1,"choice2":q.choice2,"choice3":q.choice3,"choice4":q.choice4,"choice5":q.choice5,"choice6":q.choice6,"choice7":q.choice7})
                    elif q.q_format=="openended":
                        dat.append({"format":"openended","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2})
                    else:
                        dat.append({"format":"radio","pk":q.pk,"text":q.text,"hint":q.choice1,"clue":q.choice2})

                eqno = qlist.count()+r

            else:
                return JsonResponse({"data":-1,"qno":-1,"totq":-1,"rqno":-1})
        
        return JsonResponse({"data":dat[ud.q_no-1],"qno":ud.q_no,"totq":len(dat),"rqno":r,"eqno":eqno})











@login_required(login_url='/')
def ans_ques(request):
    
    if request.method=='POST':
        ud = Userdata.objects.get(user_id=request.user)
        ud.q_no+=1
        ud.save()

        data = json.loads(request.body.decode('utf-8'))
        Answer.objects.create(q_no=Question.objects.get(pk=data['pk']),userdata_id=ud,ans=data['ans'])
        next = -1

        if ud.status == 1:
            if ud.q_no == Question.objects.filter(q_category=ud.category,q_type='prepos').count()+1:
                ud.category+=1
                ud.q_no = 1
                ud.save()
                next = 1

            if ud.category == 5:
                iid = ud.pk%10
                if iid>0 and iid<5:
                    ud.status = 2  #experiment
                    ud.group = "experiment"
                elif iid>4 and iid<9:
                    ud.status = 3  #placebo
                    ud.group = "placebo"
                else:
                    ud.status = 4  #control
                    ud.group = "control"
                ud.email_date = dt.date.today()
                ud.category = 1
                ud.save()
                next = 1

        if ud.status == 2:
            if ud.category !=1:
                if ud.q_no == 1+Question.objects.filter(q_category=ud.category,q_type='experiment').count()+Question.objects.filter(q_category=ud.category-1,q_type='experiment',q_format="mcq").count():
                    ud.category+=1
                    ud.email_date += dt.timedelta(days=7)
                    ud.q_no = 1
                    ud.save()
                    next = 1

                if ud.category == 5:
                    ud.status = 5
                    ud.category = 1
                    ud.save()
                    next = 1
            else:
                if ud.q_no == Question.objects.filter(q_category=ud.category,q_type='experiment').count()+1:
                    ud.category+=1
                    ud.email_date += dt.timedelta(days=7)
                    ud.q_no = 1
                    ud.save()
                    next = 1

                if ud.category == 5:
                    ud.status = 5
                    ud.category = 1
                    ud.save()
                    next = 1

        if ud.status == 3:
            if ud.q_no == Question.objects.filter(q_category=ud.category,q_type='placebo').count()+1:
                ud.category+=1
                ud.email_date += dt.timedelta(days=7)
                ud.q_no = 1
                ud.save()
                next = 1

            if ud.category == 5:
                ud.status = 5
                ud.category = 1
                ud.save()
                next = 1


        if ud.status == 5:
            if ud.group == "experiment" and ud.category == 1:
                if ud.q_no == 1+Question.objects.filter(q_category=ud.category,q_type='prepos').count()+Question.objects.filter(q_category=4,q_type='experiment',q_format="mcq").count():
                    ud.category+=1
                    ud.q_no = 1
                    ud.save()
                    next = 1

                if ud.category == 5:
                    ud.status = 6
                    ud.save()
                    next = 1
            else:
                if ud.q_no == Question.objects.filter(q_category=ud.category,q_type='prepos').count()+1:
                    ud.category+=1
                    ud.q_no = 1
                    ud.save()
                    next = 1

                if ud.category == 5:
                    ud.status = 6
                    ud.save()
                    next = 1

        return JsonResponse({"next":next})

@login_required(login_url='/')
def logout_view(request):
    logout(request)
    return HttpResponseRedirect('/')


@login_required(login_url='/')
def admin_page(request):
    return render(request, 'admin_panel.html')


@login_required(login_url='/')
def admin_page_details(request):
    data = json.loads(request.body.decode('utf-8'))
    q = Question()
    q.text = data['text']
    q.q_type = data['q_type']
    q.q_format = data['q_format']
    q.q_category = data['q_category']

    try:
        q.choice1 = data['choice1']
    except KeyError:
        pass

    try:
        q.choice2 = data['choice2']
    except KeyError:
        pass

    try:
        q.choice3 = data['choice3']
    except KeyError:
        pass

    try:
        q.choice4 = data['choice4']
    except KeyError:
        pass

    try:
        q.choice5 = data['choice5']
    except KeyError:
        pass

    try:
        q.choice6 = data['choice6']
    except KeyError:
        pass

    try:
        q.choice7 = data['choice7']
    except KeyError:
        pass

    q.save()
    return JsonResponse({"success":1})


def admin_db_1(request):
    queryset= Userdata.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=database.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'prepos'

    row_num = 0


    columns = [
        (u"ID", 15),
        (u"Name", 40),
        # (u"Email", 50),
        (u"Age", 50),
        (u"Gender", 50),
        (u"Education_Level", 50),
        (u"Phone", 20),

    ]
    for q in Question.objects.filter(q_type='prepos'):
        columns+=(q.text,200),


    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for ud in queryset:
        # use = User.objects.get(pk=ud.user_id)
        row_num += 1
        row = [
            ud.pk,
            ud.name,
            # use.email,
            ud.age,
            ud.gender,
            ud.education_level,
            ud.contactno,
        ]

        for an in Answer.objects.filter(userdata_id=ud, q_no__q_type='prepos'):
            row+=an.ans


        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
                #c.style.alignment.wrap_text = True

    wb.save(response)
    return response

def admin_db_2(request):
    queryset= Userdata.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=database.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'experiment'

    row_num = 0


    columns = [
        (u"ID", 15),
        (u"Name", 40),
        # (u"Email", 50),
        (u"Age", 50),
        (u"Gender", 50),
        (u"Education_Level", 50),
        (u"Phone", 20),

    ]
    for q in Question.objects.filter(q_type='experiment'):
        columns+=(q.text,200),


    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for ud in queryset:
        # use = User.objects.get(pk=ud.user_id)
        row_num += 1
        row = [
            ud.pk,
            ud.name,
            # use.email,
            ud.age,
            ud.gender,
            ud.education_level,
            ud.contactno,
        ]

        for an in Answer.objects.filter(userdata_id=ud, q_no__q_type='experiment'):
            row+=an.ans


        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
                #c.style.alignment.wrap_text = True

    wb.save(response)
    return response

def admin_db_3(request):
    queryset= Userdata.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=database.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'placebo'

    row_num = 0


    columns = [
        (u"ID", 15),
        (u"Name", 40),
        # (u"Email", 50),
        (u"Age", 50),
        (u"Gender", 50),
        (u"Education_Level", 50),
        (u"Phone", 20),

    ]
    for q in Question.objects.filter(q_type='placebo'):
        columns+=(q.text,200),


    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for ud in queryset:
        row_num += 1
        row = [
            ud.pk,
            ud.name,
            # User.objects.get(pk=ud.user_id).email,
            ud.age,
            ud.gender,
            ud.education_level,
            ud.contactno,
        ]

        for an in Answer.objects.filter(userdata_id=ud, q_no__q_type='placebo'):
            row+=an.ans


        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
                #c.style.alignment.wrap_text = True

    wb.save(response)
    return response

def admin_db_4(request):
    queryset= Userdata.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=database.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'control'

    row_num = 0


    columns = [
        (u"ID", 15),
        (u"Name", 40),
        # (u"Email", 50),
        (u"Age", 50),
        (u"Gender", 50),
        (u"Education_Level", 50),
        (u"Phone", 20),

    ]
    for q in Question.objects.filter(q_type='control'):
        columns+=(q.text,200),


    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for ud in queryset:
        row_num += 1
        row = [
            ud.pk,
            ud.name,
            # User.objects.get(pk=ud.user_id).email,
            ud.age,
            ud.gender,
            ud.education_level,
            ud.contactno,
        ]

        for an in Answer.objects.filter(userdata_id=ud, q_no__q_type='control'):
            row+=an.ans


        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
                #c.style.alignment.wrap_text = True

    wb.save(response)
    return response
